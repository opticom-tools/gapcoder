# -------------------------------------------------------------
# GapCoder v1.5 — GAP Analysis Summariser (Importance / Client / Competitor + Comments)
# Last updated: 2026-01-15
#
# Supported headers (ONLY; no fallback):
#   gap_imp_1, gap_perf_1, gap_comp_1, gap_comm_1
#
# Input:
# - Upload .xlsx (recommended when comments are long / may contain line breaks)
# - Paste TSV/CSV copied from Excel (works when rows are clean)
#
# Gap Dictionary (required):
#   1 = Section Name | Criterion Name
#
# Missing rules (excluded from all means & gaps):
# - 999 = don't know
# - 0   = no answer
# - blank = missing
#
# IMPORTANT (pairwise gaps):
# - Gap vs Expectations is computed per respondent only when BOTH Importance and Client exist:
#     (Importance - Client)
# - Gap vs Competitor is computed per respondent only when BOTH Client and Competitor exist:
#     (Client - Competitor)
# - Means of gaps are averages of those respondent-level gaps (NOT gap-of-means).
#
# Competitive position buckets (based on Gap vs Competitor = Client - Competitor):
# - "competitive advantages" if > +0.30
# - "behind competition" if < -0.30
# - "similar to competition" otherwise
#
# Claude output reliability:
# - Uses Anthropic tool calling to force structured output (prevents invalid JSON).
# -------------------------------------------------------------

import streamlit as st
import os
import json
import re
import csv
from io import StringIO, BytesIO
from datetime import datetime
from anthropic import Anthropic

# Excel support
try:
    from openpyxl import load_workbook, Workbook
    XLSX_OK = True
except Exception:
    XLSX_OK = False

# PDF support
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    pdf_supported = True
except Exception:
    pdf_supported = False


# --------------------
# Config
# --------------------
CLAUDE_MODEL = "claude-sonnet-4-5"
MAX_CLAUDE_TOKENS = 2600
PROJECTS_FILE = "gapcoder_projects.json"

# threshold for competitive position buckets (Client - Competitor)
COMP_TOL = 0.30

# Supported: gap_imp_1, gap_perf_1, gap_comp_1, gap_comm_1
GAP_COL_RE = re.compile(r"^gap_(imp|perf|comp|comm)_(\d+)$", re.IGNORECASE)

RESP_ID_CANDIDATES = {"resp_id", "respondent_id", "respondent", "resp", "id"}
ID_LIKE_RE = re.compile(r"^[A-Za-z]{2,}\d+.*$")  # e.g. UPM001, RESP_001, etc.


# --------------------
# Project persistence
# --------------------
def load_projects(path: str) -> dict:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_projects(path: str, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# --------------------
# Parsing helpers
# --------------------
def detect_delimiter_from_header(header_line: str) -> str:
    if "\t" in header_line:
        return "\t"
    if ";" in header_line and "," not in header_line:
        return ";"
    return ","

def normalise_headers(fieldnames):
    """Replace empty headers with COL_1, COL_2 etc., and make them unique."""
    seen = set()
    out = []
    for idx, h in enumerate(fieldnames):
        name = "" if h is None else str(h).strip()
        if name == "":
            name = f"COL_{idx+1}"
        base = name
        i = 2
        while name in seen:
            name = f"{base}_{i}"
            i += 1
        seen.add(name)
        out.append(name)
    return out

def parse_table_paste(text: str):
    if not text or not text.strip():
        raise ValueError("No data pasted.")

    # Remove fully empty lines
    lines = [ln for ln in text.splitlines() if ln.strip() != ""]
    if len(lines) < 2:
        raise ValueError(
            "It looks like you only pasted one line (often just the header, or header+values without row breaks). "
            "Try copying the full Excel table again, or use the .xlsx upload option."
        )

    delim = detect_delimiter_from_header(lines[0])
    cleaned_text = "\n".join(lines)
    f = StringIO(cleaned_text)
    reader = csv.reader(f, delimiter=delim)

    raw_fieldnames = next(reader, None)
    if not raw_fieldnames:
        raise ValueError("Could not read header row. Make sure the first line is the header.")

    headers = normalise_headers(raw_fieldnames)

    rows = []
    for row in reader:
        # Skip empty rows
        if not any(str(v).strip() for v in row if v is not None):
            continue
        # Pad / trim to header length
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        if len(row) > len(headers):
            row = row[:len(headers)]
        rows.append({headers[i]: ("" if row[i] is None else str(row[i]).strip()) for i in range(len(headers))})

    if not rows:
        raise ValueError(
            "No rows found under the header. This can happen when comments contain hidden line breaks, "
            "or the paste didn't include respondent rows. Recommendation: upload the .xlsx instead."
        )

    return headers, rows

def parse_table_xlsx(uploaded_file):
    if not XLSX_OK:
        raise ValueError("openpyxl is not available. Add 'openpyxl' to requirements.txt.")

    data = uploaded_file.getvalue()
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active

    # Header row
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if not raw_headers or all(h is None or str(h).strip() == "" for h in raw_headers):
        raise ValueError("Row 1 in the Excel file is empty. Please ensure row 1 contains headers.")

    headers = normalise_headers(raw_headers)

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None:
            continue
        if not any(v is not None and str(v).strip() != "" for v in r):
            continue

        row = list(r)
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        if len(row) > len(headers):
            row = row[:len(headers)]

        cleaned = {}
        for i, h in enumerate(headers):
            v = row[i]
            cleaned[h] = "" if v is None else v
        rows.append(cleaned)

    if not rows:
        raise ValueError("No data rows found under the header in the Excel file.")
    return headers, rows

def parse_score(value):
    """Returns numeric float or None. Treats 0, 999 and blanks as missing."""
    if value is None:
        return None

    # Numeric already (xlsx path)
    if isinstance(value, (int, float)):
        num = float(value)
    else:
        s = str(value).strip()
        if s == "":
            return None
        try:
            num = float(s.replace(",", "."))
        except:
            return None

    if abs(num - 0.0) < 1e-12:
        return None
    if abs(num - 999.0) < 1e-12:
        return None
    return num

def parse_gap_dictionary(text: str):
    """
    Expected:
      12 = Section Name | Criterion Name
    Also accepts ":" instead of "=".
    """
    mapping = {}
    if not text or not text.strip():
        return mapping

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if "=" in line:
            left, right = line.split("=", 1)
        elif ":" in line:
            left, right = line.split(":", 1)
        else:
            continue

        try:
            gap_no = int(left.strip())
        except:
            continue

        right = right.strip()
        if "|" in right:
            section, crit = right.split("|", 1)
            section = section.strip() or "General"
            crit = crit.strip() or f"Gap {gap_no}"
        else:
            section = "General"
            crit = right.strip() or f"Gap {gap_no}"

        mapping[gap_no] = {"section": section, "criterion": crit}

    return mapping

def build_gap_schema(headers):
    """
    Returns:
      gaps: sorted list of gap numbers
      col_map: dict[(gap_no, suffix)] -> header
    suffix: imp, perf, comp, comm
    """
    col_map = {}
    gap_nums = set()

    for h in headers:
        hh = str(h).strip()
        m = GAP_COL_RE.match(hh)
        if not m:
            continue
        suffix = m.group(1).lower()
        n = int(m.group(2))
        col_map[(n, suffix)] = h
        gap_nums.add(n)

    return sorted(gap_nums), col_map

def detect_resp_id_col(headers, rows):
    # 1) by name
    for h in headers:
        if str(h).strip().lower() in RESP_ID_CANDIDATES:
            return h

    # 2) heuristic: look for a column that is mostly "id-like"
    best = None
    best_score = 0.0
    for h in headers:
        vals = []
        for r in rows[:50]:
            vals.append(r.get(h, ""))
        nonempty = [v for v in vals if str(v).strip() != ""]
        if len(nonempty) < 5:
            continue
        idlike = sum(1 for v in nonempty if ID_LIKE_RE.match(str(v).strip()))
        score = idlike / max(1, len(nonempty))
        if score > best_score:
            best_score = score
            best = h

    if best_score >= 0.6:
        return best
    return None

def mean(vals):
    vv = [v for v in vals if v is not None]
    if not vv:
        return None
    return sum(vv) / len(vv)

def mean_gap(vals):
    """Mean of respondent-level gaps (already computed), excluding None."""
    vv = [v for v in vals if v is not None]
    if not vv:
        return None
    return sum(vv) / len(vv)

def classify_competitive_position(gap_client_minus_competitor):
    """Buckets based on (Client - Competitor)."""
    if gap_client_minus_competitor is None:
        return None
    if gap_client_minus_competitor > COMP_TOL:
        return "competitive_advantages"
    if gap_client_minus_competitor < -COMP_TOL:
        return "behind_competition"
    return "similar_to_competition"

def compute_tables(rows, gaps, col_map, resp_id_col, gap_dict):
    """
    Returns:
      criterion_table: list[dict]
      comments_by_gap: dict gap_no -> list of {"resp_id","comment"} (truncated)
    """
    criterion_table = []
    comments_by_gap = {}

    # Collect comments by gap
    for n in gaps:
        comm_col = col_map.get((n, "comm"))
        if not comm_col:
            continue
        bucket = []
        for i, r in enumerate(rows, start=1):
            resp_id = (r.get(resp_id_col) if resp_id_col else None)
            resp_id = str(resp_id).strip() if resp_id is not None and str(resp_id).strip() != "" else f"RESP_{i:03d}"
            c = r.get(comm_col, "")
            c = "" if c is None else str(c).strip()
            if c:
                if len(c) > 320:
                    c = c[:320].rstrip() + "…"
                bucket.append({"resp_id": resp_id, "comment": c})
        if bucket:
            comments_by_gap[n] = bucket

    # Compute means + pairwise gaps
    for n in gaps:
        imp_col = col_map.get((n, "imp"))
        perf_col = col_map.get((n, "perf"))
        comp_col = col_map.get((n, "comp"))

        imp_vals, perf_vals, comp_vals = [], [], []
        gap_vs_expect_vals = []  # (imp - perf) per respondent if both exist
        gap_vs_comp_vals = []    # (perf - comp) per respondent if both exist

        for r in rows:
            imp = parse_score(r.get(imp_col, "") if imp_col else "")
            perf = parse_score(r.get(perf_col, "") if perf_col else "")
            comp = parse_score(r.get(comp_col, "") if comp_col else "")

            imp_vals.append(imp)
            perf_vals.append(perf)
            comp_vals.append(comp)

            # Pairwise expectation gap only when BOTH exist
            gap_vs_expect_vals.append((imp - perf) if (imp is not None and perf is not None) else None)

            # Pairwise competitor gap only when BOTH exist
            gap_vs_comp_vals.append((perf - comp) if (perf is not None and comp is not None) else None)

        meta = gap_dict.get(n, {"section": "Unmapped", "criterion": f"Gap {n}"})

        imp_mean = mean(imp_vals)
        perf_mean = mean(perf_vals)
        comp_mean = mean(comp_vals)

        gap_expect_mean = mean_gap(gap_vs_expect_vals)   # Importance - Client
        gap_comp_mean = mean_gap(gap_vs_comp_vals)       # Client - Competitor

        criterion_table.append({
            "gap_no": n,
            "section": meta["section"],
            "criterion": meta["criterion"],

            "mean_importance": None if imp_mean is None else round(imp_mean, 2),
            "mean_client": None if perf_mean is None else round(perf_mean, 2),
            "mean_competitor": None if comp_mean is None else round(comp_mean, 2),

            # Means of respondent-level gaps
            "mean_gap_vs_expectations_imp_minus_perf": None if gap_expect_mean is None else round(gap_expect_mean, 2),
            "mean_gap_vs_competitor_perf_minus_comp": None if gap_comp_mean is None else round(gap_comp_mean, 2),

            # Counts (useful for reliability)
            "n_importance": sum(1 for v in imp_vals if v is not None),
            "n_client": sum(1 for v in perf_vals if v is not None),
            "n_competitor": sum(1 for v in comp_vals if v is not None),
            "n_gap_vs_expectations": sum(1 for v in gap_vs_expect_vals if v is not None),
            "n_gap_vs_competitor": sum(1 for v in gap_vs_comp_vals if v is not None),
        })

    return criterion_table, comments_by_gap

def summarise(criteria_rows):
    total = len(criteria_rows)
    eval_comp = [r for r in criteria_rows if r["mean_gap_vs_competitor_perf_minus_comp"] is not None]
    eval_expect = [r for r in criteria_rows if r["mean_gap_vs_expectations_imp_minus_perf"] is not None]

    adv = sim = beh = 0
    for r in eval_comp:
        cls = classify_competitive_position(r["mean_gap_vs_competitor_perf_minus_comp"])
        if cls == "competitive_advantages":
            adv += 1
        elif cls == "behind_competition":
            beh += 1
        else:
            sim += 1

    # Top lists
    top_adv = sorted(eval_comp, key=lambda r: r["mean_gap_vs_competitor_perf_minus_comp"], reverse=True)[:3]
    top_beh = sorted(eval_comp, key=lambda r: r["mean_gap_vs_competitor_perf_minus_comp"])[:3]
    top_expect = sorted(eval_expect, key=lambda r: r["mean_gap_vs_expectations_imp_minus_perf"], reverse=True)[:3]

    avg_imp = mean([r["mean_importance"] for r in criteria_rows])
    avg_cli = mean([r["mean_client"] for r in criteria_rows])
    avg_com = mean([r["mean_competitor"] for r in criteria_rows])
    avg_gap_comp = mean([r["mean_gap_vs_competitor_perf_minus_comp"] for r in eval_comp])
    avg_gap_expect = mean([r["mean_gap_vs_expectations_imp_minus_perf"] for r in eval_expect])

    return {
        "criteria_total": total,
        "criteria_evaluable_vs_competitor": len(eval_comp),
        "criteria_evaluable_vs_expectations": len(eval_expect),

        "competitive_position_counts_based_on_client_minus_competitor": {
            "competitive_advantages": adv,
            "similar_to_competition": sim,
            "behind_competition": beh,
            "threshold": COMP_TOL
        },

        "averages": {
            "importance": None if avg_imp is None else round(avg_imp, 2),
            "client": None if avg_cli is None else round(avg_cli, 2),
            "competitor": None if avg_com is None else round(avg_com, 2),
            "gap_vs_competitor_client_minus_competitor": None if avg_gap_comp is None else round(avg_gap_comp, 2),
            "gap_vs_expectations_importance_minus_client": None if avg_gap_expect is None else round(avg_gap_expect, 2)
        },

        "top3_competitive_advantages_client_minus_competitor": [
            {"gap_no": r["gap_no"], "section": r["section"], "criterion": r["criterion"],
             "gap_client_minus_competitor": r["mean_gap_vs_competitor_perf_minus_comp"]}
            for r in top_adv
        ],
        "top3_behind_competition_client_minus_competitor": [
            {"gap_no": r["gap_no"], "section": r["section"], "criterion": r["criterion"],
             "gap_client_minus_competitor": r["mean_gap_vs_competitor_perf_minus_comp"]}
            for r in top_beh
        ],
        "top3_gaps_vs_expectations_importance_minus_client": [
            {"gap_no": r["gap_no"], "section": r["section"], "criterion": r["criterion"],
             "gap_importance_minus_client": r["mean_gap_vs_expectations_imp_minus_perf"]}
            for r in top_expect
        ],

        "note": "Missing: 0 / 999 / blanks excluded. Gaps are pairwise (both values must be present)."
    }

def pick_comment_samples(criteria_table, comments_by_gap, max_gaps=12, per_gap=6):
    """Keep prompt small: focus on biggest expectation gaps and most negative competitor gaps."""
    eval_expect = [r for r in criteria_table if r["mean_gap_vs_expectations_imp_minus_perf"] is not None]
    eval_comp = [r for r in criteria_table if r["mean_gap_vs_competitor_perf_minus_comp"] is not None]

    top_expect = sorted(eval_expect, key=lambda r: r["mean_gap_vs_expectations_imp_minus_perf"], reverse=True)[:6]
    top_beh = sorted(eval_comp, key=lambda r: r["mean_gap_vs_competitor_perf_minus_comp"])[:6]  # most negative first

    priority = []
    seen = set()
    for r in top_expect + top_beh:
        n = r["gap_no"]
        if n not in seen:
            priority.append(n)
            seen.add(n)
    priority = priority[:max_gaps]

    out = {}
    for n in priority:
        bucket = comments_by_gap.get(n, [])
        if bucket:
            out[n] = bucket[:per_gap]
    return out

def to_csv_bytes(rows):
    if not rows:
        return b""
    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")

def to_xlsx_bytes(rows):
    if not XLSX_OK or not rows:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Criterion table"
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def build_pdf(ctx, parsed, criteria_rows_slim=None):
    """Create a simple PDF report in the OptiCoder spirit (cover + sections)."""
    if not pdf_supported:
        return None

    styles = getSampleStyleSheet()
    normal = styles["BodyText"]
    heading = styles["Heading2"]
    h1 = styles["Heading1"]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="GapCoder Report"
    )
    elems = []

    # Cover
    elems.append(Paragraph("GapCoder – GAP Analysis Summary", h1))
    elems.append(Spacer(1, 4 * mm))
    meta = (
        f"Project: {ctx.get('project_no','')}<br/>"
        f"Client: {ctx.get('client_name','')}<br/>"
        f"Industry: {ctx.get('industry','')}<br/>"
        f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
    )
    elems.append(Paragraph(meta, normal))
    elems.append(Spacer(1, 4 * mm))
    defs = (
        "<b>Definitions</b><br/>"
        "Gap vs Expectations = (Importance − Client). Positive means under-delivery vs what matters.<br/>"
        "Gap vs Competitor = (Client − Competitor). Positive means client leads competitor.<br/>"
        "Missing values (0 / 999 / blanks) are excluded. Gaps are pairwise (both values must exist)."
    )
    elems.append(Paragraph(defs, normal))
    elems.append(PageBreak())

    # Total
    total = parsed.get("total_gap_overview", {})
    elems.append(Paragraph("Total GAP overview", heading))
    elems.append(Spacer(1, 2 * mm))
    bullets = total.get("slide_bullets", []) or []
    if bullets:
        for b in bullets:
            elems.append(Paragraph(f"• {str(b)}", normal))
        elems.append(Spacer(1, 4 * mm))
    narrative = total.get("narrative", "") or ""
    if narrative.strip():
        elems.append(Paragraph(narrative.replace("\n", "<br/>"), normal))
    elems.append(PageBreak())

    # Sections
    secs = parsed.get("sections", []) or []
    for s in secs:
        sec_name = s.get("section", "Unnamed section")
        elems.append(Paragraph(sec_name, heading))
        elems.append(Spacer(1, 2 * mm))
        for b in (s.get("slide_bullets", []) or []):
            elems.append(Paragraph(f"• {str(b)}", normal))
        elems.append(Spacer(1, 4 * mm))
        nar = s.get("narrative", "") or ""
        if nar.strip():
            elems.append(Paragraph(nar.replace("\n", "<br/>"), normal))
        elems.append(PageBreak())

    # Optional: small appendix note
    if criteria_rows_slim:
        elems.append(Paragraph("Appendix (excerpt): Criterion table", heading))
        elems.append(Spacer(1, 2 * mm))
        note = "Full criterion table is available as a downloadable file from the app."
        elems.append(Paragraph(note, normal))

    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()


# --------------------
# Claude: tool-based structured output
# --------------------
TOOL_SCHEMA = {
    "name": "gap_analysis",
    "description": "Return the GAP analysis in a strict structured format.",
    "input_schema": {
        "type": "object",
        "properties": {
            "total_gap_overview": {
                "type": "object",
                "properties": {
                    "slide_bullets": {"type": "array", "items": {"type": "string"}},
                    "narrative": {"type": "string"}
                },
                "required": ["slide_bullets", "narrative"]
            },
            "sections": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "section": {"type": "string"},
                        "slide_bullets": {"type": "array", "items": {"type": "string"}},
                        "narrative": {"type": "string"}
                    },
                    "required": ["section", "slide_bullets", "narrative"]
                }
            }
        },
        "required": ["total_gap_overview", "sections"]
    }
}

def build_prompt(ctx, criteria_for_prompt, overall_stats, section_stats, comment_samples):
    return f"""
You are a senior market research consultant. Write a GAP analysis summary that is slide-ready and grounded ONLY in the provided statistics and comment samples.

DEFINITIONS (use these labels consistently):
- Gap vs Expectations = (Importance - Client). Positive means: under-delivery vs what matters.
- Gap vs Competitor = (Client - Competitor). Positive means: client leads competitor.

COMPETITIVE POSITION BUCKETS (based on Gap vs Competitor = Client - Competitor):
- "Competitive advantages" if > +0.30
- "Similar to competition" if between -0.30 and +0.30
- "Behind competition" if < -0.30

RULES:
- Do NOT invent numbers.
- Use the provided overall/section stats and criterion table.
- Missing handling: 0, 999 and blanks are excluded. Gaps are computed pairwise (both values must exist).
- Be explicit when referring to a gap: always state whether it is Gap vs Expectations (Importance - Client) or Gap vs Competitor (Client - Competitor).

PROJECT CONTEXT:
- Project: {ctx.get("project_no","")}
- Client: {ctx.get("client_name","")}
- Industry: {ctx.get("industry","")}
- Objectives: {ctx.get("objectives","")}

OVERALL STATS:
{json.dumps(overall_stats, ensure_ascii=False, indent=2)}

SECTION STATS:
{json.dumps(section_stats, ensure_ascii=False, indent=2)}

CRITERION TABLE:
{json.dumps(criteria_for_prompt, ensure_ascii=False, indent=2)}

COMMENT SAMPLES (use for improvement suggestions; refer to section/criterion when possible):
{json.dumps(comment_samples, ensure_ascii=False, indent=2)}

WHAT TO PRODUCE (as tool output):
1) Total GAP overview (200–400 words narrative + slide bullets)
   Include:
   - how many criteria evaluated
   - high-level remark
   - competitive advantages / similar / behind (where evaluable, based on Gap vs Competitor)
   - top 3 competitive advantages (Client - Competitor) and weakest positions (smallest/negative Client - Competitor)
   - largest gaps vs expectations (Importance - Client)
   - key improvement themes based on comment samples
2) Key findings per section (200–300 words each + slide bullets)
   Cover:
   - expectations met in general (Importance vs Client)
   - top gaps vs expectations (Importance - Client)
   - performance vs competition (Client - Competitor)
   - weakest positions vs competition
   - improvement suggestions based on comment samples
""".strip()

def extract_tool_output(msg, tool_name="gap_analysis"):
    """Return the tool input dict if present, else None."""
    content = getattr(msg, "content", None)
    if not content:
        return None
    for block in content:
        btype = getattr(block, "type", None)
        name = getattr(block, "name", None)
        if btype == "tool_use" and name == tool_name:
            return getattr(block, "input", None)
    return None


# --------------------
# Streamlit UI
# --------------------
st.set_page_config(page_title="GapCoder", layout="wide")
st.markdown(f"# 📊 GapCoder (v1.5)\n_Last updated: {datetime.now():%Y-%m-%d}_")

projects = load_projects(PROJECTS_FILE)
client = Anthropic(api_key=st.secrets.get("ANTHROPIC_API_KEY", ""))

with st.expander("1. Project Context", expanded=True):
    sel = st.selectbox("Load project:", ["-- New --"] + list(projects.keys()))
    project_no = sel if sel != "-- New --" else st.text_input("Project Number")
    defaults = projects.get(project_no, {}) if project_no else {}

    c1, c2 = st.columns(2)
    with c1:
        client_name = st.text_input("Client Name", value=defaults.get("client_name", ""))
        industry = st.text_input("Industry", value=defaults.get("industry", ""))
        mode = st.radio("Analysis mode", ["All sections", "One section"], index=0)
    with c2:
        objectives = st.text_area("Project Objectives", value=defaults.get("objectives", ""), height=90)

    gap_dict_raw = st.text_area(
        "Gap Dictionary (required): gap number = Section | Gap name",
        value=defaults.get("gap_dict_raw", ""),
        height=180,
        help="Example:\n1 = Communication | Proactive updates\n2 = Communication | Clarity of information\n6 = Supply Chain | Delivery reliability"
    )

    ctx = {
        "project_no": project_no,
        "client_name": client_name,
        "industry": industry,
        "objectives": objectives,
        "mode": mode,
        "gap_dict_raw": gap_dict_raw
    }

    if project_no:
        projects[project_no] = ctx
        save_projects(PROJECTS_FILE, projects)

    gd = parse_gap_dictionary(gap_dict_raw)
    if gd:
        secs = sorted({v["section"] for v in gd.values()})
        st.info(f"✅ Gap Dictionary loaded: {len(gd)} gaps mapped across {len(secs)} sections.")
    else:
        st.warning("Gap Dictionary is required before running the analysis.")

with st.expander("2. Input data (simple)", expanded=True):
    st.markdown(
        "**Recommended:** Upload the Excel file (.xlsx). This avoids copy/paste issues when comments contain line breaks.\n\n"
        "**Supported headers:** `gap_imp_1`, `gap_perf_1`, `gap_comp_1`, `gap_comm_1` (and so on).\n\n"
        "Missing rules: blank / `0` / `999` are treated as missing and ignored in all means and gaps."
    )

    input_mode = st.radio("How do you want to provide data?", ["Upload Excel (.xlsx)", "Paste table (TSV/CSV)"], index=0)

    uploaded = None
    raw_text = ""

    if input_mode == "Upload Excel (.xlsx)":
        if not XLSX_OK:
            st.error("Excel upload needs openpyxl. Add it to requirements.txt and redeploy.")
        uploaded = st.file_uploader("Upload .xlsx", type=["xlsx"])
        st.caption("Tip: row 1 = headers. Each following row = one respondent.")
    else:
        st.code(
            "ID\tgap_imp_1\tgap_perf_1\tgap_comp_1\tgap_comm_1\n"
            "UPM001\t7\t8\t7\t(optional comment)\n"
            "UPM002\t9\t9\t\tRefused competitor rating (leave competitor blank)\n",
            language="text"
        )
        raw_text = st.text_area("Paste table here", height=260, key="raw_gap")

# Session outputs (for downloads)
if "gapcoder_parsed" not in st.session_state:
    st.session_state["gapcoder_parsed"] = None
if "gapcoder_criteria" not in st.session_state:
    st.session_state["gapcoder_criteria"] = None
if "gapcoder_ctx" not in st.session_state:
    st.session_state["gapcoder_ctx"] = None

if st.button("🧠 Generate GAP Analysis"):
    if not st.secrets.get("ANTHROPIC_API_KEY"):
        st.error("Missing ANTHROPIC_API_KEY in Streamlit Secrets.")
        st.stop()

    if not ctx.get("gap_dict_raw", "").strip():
        st.error("Please fill in the Gap Dictionary first (gap number = Section | Gap name).")
        st.stop()

    gap_dict = parse_gap_dictionary(ctx["gap_dict_raw"])
    if not gap_dict:
        st.error("Gap Dictionary could not be parsed. Please use: 1 = Section | Gap name")
        st.stop()

    # Parse data
    try:
        if input_mode == "Upload Excel (.xlsx)":
            if uploaded is None:
                st.error("Please upload an .xlsx file first.")
                st.stop()
            headers, rows = parse_table_xlsx(uploaded)
        else:
            if not raw_text.strip():
                st.error("Paste your GAP table first (header + rows).")
                st.stop()
            headers, rows = parse_table_paste(raw_text)
    except ValueError as e:
        st.error(str(e))
        st.stop()

    gaps, col_map = build_gap_schema(headers)
    if not gaps:
        st.error("Could not find any gap columns. Expected headers like gap_imp_1 / gap_perf_1 / gap_comp_1 / gap_comm_1.")
        st.stop()

    # Ensure dictionary covers all gaps found
    missing_in_dict = [n for n in gaps if n not in gap_dict]
    if missing_in_dict:
        st.error(f"Gap Dictionary is missing these gap numbers: {missing_in_dict}. Please add them.")
        st.stop()

    resp_id_col = detect_resp_id_col(headers, rows)
    if resp_id_col is None:
        st.warning("No RESP_ID/ID column detected. Rows will be labelled as RESP_001, RESP_002, ...")

    # Filter for one section (optional)
    selected_section = None
    if ctx.get("mode") == "One section":
        all_sections = sorted({gap_dict[n]["section"] for n in gaps})
        selected_section = st.selectbox("Choose section to analyse", all_sections, index=0)
        gaps = [n for n in gaps if gap_dict[n]["section"] == selected_section]
        if not gaps:
            st.error("No gaps found for selected section.")
            st.stop()

    # Compute tables
    criterion_table, comments_by_gap = compute_tables(
        rows=rows,
        gaps=gaps,
        col_map=col_map,
        resp_id_col=resp_id_col,
        gap_dict=gap_dict
    )

    overall_stats = summarise(criterion_table)

    grouped = {}
    for r in criterion_table:
        grouped.setdefault(r["section"], []).append(r)
    section_stats = {sec: summarise(items) for sec, items in grouped.items()}

    comment_samples = pick_comment_samples(criterion_table, comments_by_gap, max_gaps=12, per_gap=6)

    # Reduce table for Claude (smaller + explicit gap names)
    criteria_for_prompt = []
    for r in criterion_table:
        criteria_for_prompt.append({
            "gap_no": r["gap_no"],
            "section": r["section"],
            "criterion": r["criterion"],

            "mean_importance": r["mean_importance"],
            "mean_client": r["mean_client"],
            "mean_competitor": r["mean_competitor"],

            "mean_gap_vs_expectations_importance_minus_client": r["mean_gap_vs_expectations_imp_minus_perf"],
            "mean_gap_vs_competitor_client_minus_competitor": r["mean_gap_vs_competitor_perf_minus_comp"],

            "n_gap_vs_expectations": r["n_gap_vs_expectations"],
            "n_gap_vs_competitor": r["n_gap_vs_competitor"],
        })

    st.subheader("Quick sanity check (computed from your data)")
    st.write(overall_stats)

    st.markdown("### Criterion table (means & gaps)")
    st.dataframe(criteria_for_prompt, use_container_width=True)

    prompt = build_prompt(ctx, criteria_for_prompt, overall_stats, section_stats, comment_samples)

    with st.spinner("🤖 Claude is thinking…"):
        msg = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=MAX_CLAUDE_TOKENS,
            temperature=0.2,
            messages=[{"role": "user", "content": prompt}],
            tools=[TOOL_SCHEMA],
            tool_choice={"type": "tool", "name": "gap_analysis"},
        )

    parsed = extract_tool_output(msg, tool_name="gap_analysis")

    if not isinstance(parsed, dict):
        st.error("Claude did not return the structured tool output. Showing raw output below.")
        # Best-effort raw text display:
        raw_text_out = ""
        try:
            content = getattr(msg, "content", [])
            for b in content:
                if getattr(b, "type", "") == "text":
                    raw_text_out += getattr(b, "text", "") + "\n"
        except Exception:
            pass
        st.text_area("Raw output", raw_text_out or "(no text returned)", height=320)
        st.stop()

    # Store for downloads
    st.session_state["gapcoder_parsed"] = parsed
    st.session_state["gapcoder_criteria"] = criteria_for_prompt
    st.session_state["gapcoder_ctx"] = ctx

# --------------------
# Display outputs + downloads (if available)
# --------------------
parsed = st.session_state.get("gapcoder_parsed")
criteria_for_prompt = st.session_state.get("gapcoder_criteria")
ctx_saved = st.session_state.get("gapcoder_ctx")

if parsed and criteria_for_prompt and ctx_saved:
    total = parsed.get("total_gap_overview", {})
    sections_out = parsed.get("sections", [])

    tabs = st.tabs(["Total overview", "By section", "Downloads", "Copy"])

    with tabs[0]:
        st.markdown("### Slide bullets")
        st.text_area("Bullets", "\n".join(total.get("slide_bullets", []) or []), height=190)
        st.markdown("### Narrative")
        st.text_area("Narrative", total.get("narrative", "") or "", height=280)

    with tabs[1]:
        if not sections_out:
            st.info("No section outputs returned.")
        else:
            names = [s.get("section", "Unnamed") for s in sections_out]
            pick = st.selectbox("Choose section output:", names, index=0)
            chosen = next((s for s in sections_out if s.get("section") == pick), sections_out[0])

            st.markdown("### Slide bullets")
            st.text_area("Bullets", "\n".join(chosen.get("slide_bullets", []) or []), height=190)
            st.markdown("### Narrative")
            st.text_area("Narrative", chosen.get("narrative", "") or "", height=280)

    with tabs[2]:
        st.markdown("### Download criterion table")
        csv_bytes = to_csv_bytes(criteria_for_prompt)
        st.download_button(
            "⬇️ Download criterion table (CSV)",
            data=csv_bytes,
            file_name=f"GapCoder_CriterionTable_{ctx_saved.get('project_no','project')}.csv",
            mime="text/csv"
        )

        xlsx_bytes = to_xlsx_bytes(criteria_for_prompt)
        if xlsx_bytes:
            st.download_button(
                "⬇️ Download criterion table (Excel)",
                data=xlsx_bytes,
                file_name=f"GapCoder_CriterionTable_{ctx_saved.get('project_no','project')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("### Download PDF report")
        if pdf_supported:
            pdf_bytes = build_pdf(ctx_saved, parsed, criteria_rows_slim=criteria_for_prompt)
            st.download_button(
                "⬇️ Download PDF",
                data=pdf_bytes,
                file_name=f"GapCoder_Report_{ctx_saved.get('project_no','project')}.pdf",
                mime="application/pdf"
            )
        else:
            st.info("PDF export unavailable (reportlab not installed).")

    with tabs[3]:
        st.markdown("### Total overview (copy)")
        st.text_area(
            "Total (copy)",
            "BULLETS:\n" + "\n".join(total.get("slide_bullets", []) or []) +
            "\n\nNARRATIVE:\n" + (total.get("narrative", "") or ""),
            height=290
        )

        st.markdown("### Sections (copy)")
        combined = []
        for s in sections_out:
            combined.append(f"== {s.get('section','Unnamed')} ==\n")
            combined.append("BULLETS:\n" + "\n".join(s.get("slide_bullets", []) or []) + "\n")
            combined.append("NARRATIVE:\n" + (s.get("narrative", "") or "") + "\n\n")
        st.text_area("Sections (copy)", "\n".join(combined), height=380)

# Sidebar context
st.sidebar.header("Project Context")
for k, v in ctx.items():
    if k.endswith("_raw"):
        st.sidebar.markdown(f"**{k.replace('_',' ').title()}:** (configured)")
    else:
        st.sidebar.markdown(f"**{k.replace('_',' ').title()}:** {v}")
